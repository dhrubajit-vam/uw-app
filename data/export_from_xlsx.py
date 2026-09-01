"""
Export Scored_Dataset_Full from the source Excel workbook into a clean,
typed pandas DataFrame saved as parquet. This is the ONLY script that
touches the .xlsx file - everything downstream reads the parquet.

Run once (or whenever the source workbook changes):
    python data/export_from_xlsx.py --source /path/to/workbook.xlsx
"""
import argparse
import sys
from pathlib import Path

import openpyxl
import pandas as pd

OUTPUT_PATH = Path(__file__).parent / "scored_dataset_full.parquet"


def export(source_path: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(source_path, data_only=True)
    ws = wb["Scored_Dataset_Full"]

    headers = [c.value for c in ws[2]]
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[1] is None:  # Policy_ID empty -> end of data
            continue
        rows.append(row)

    df = pd.DataFrame(rows, columns=headers)

    # Drop the label/leakage-adjacent columns we don't need at all
    df = df.drop(columns=["Data_Source_Label"], errors="ignore")

    # Coerce obvious numeric columns (openpyxl sometimes gives mixed types)
    numeric_like = [
        "Prior_Year_Premium", "Customer_Tenure_Years", "Driver_Age", "Years_Licensed",
        "Annual_Mileage", "Coverage_Lapse_Days", "Prior_Claims_3Y", "At_Fault_Claims_3Y",
        "Prior_Claims_5Y", "At_Fault_Claims_5Y", "Moving_Violations_3Y", "Speeding_Violations_3Y",
        "Months_Since_Last_Claim", "Hard_Braking_Rate", "Rapid_Acceleration_Rate",
        "Speeding_Rate", "Night_Driving_Percentage", "Distracted_Driving_Score",
        "Telematics_Safety_Score", "Vehicle_Year", "Vehicle_Age", "Vehicle_Value",
        "Vehicle_Safety_Rating", "Parts_Cost_Index", "Labor_Cost_Index", "Repairability_Score",
        "Traffic_Congestion_Score", "Vehicle_Theft_Risk_Score", "Weather_Risk_Score",
        "Hail_Risk_Score", "Flood_Risk_Score", "Hurricane_Risk_Score",
        "Litigation_Environment_Score", "Repair_Cost_Index", "Medical_Cost_Index",
        "Uninsured_Motorist_Risk", "CAT_Exposure_Score", "Collision_Deductible",
        "Comprehensive_Deductible", "PIP_Limit", "Quote_Completion_Minutes",
        "Quote_Revisions", "Agent_Engagement_Score", "Digital_Engagement_Score",
        "Final_Quoted_Premium", "Prior_Year_Premium", "Bound_Flag", "Claim_Flag",
        "Claim_Count", "Incurred_Loss", "Earned_Premium",
    ]
    for c in numeric_like:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    if "Quote_Date" in df.columns:
        df["Quote_Date"] = pd.to_datetime(df["Quote_Date"], errors="coerce")

    df.to_parquet(OUTPUT_PATH, index=False)
    print(f"Exported {len(df):,} rows x {len(df.columns)} cols -> {OUTPUT_PATH}")
    return df


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--source",
        default="/mnt/user-data/uploads/US_Personal_Auto_UW_Pricing_Prototype_dh.xlsx",
    )
    args = parser.parse_args()
    if not Path(args.source).exists():
        print(f"Source file not found: {args.source}", file=sys.stderr)
        sys.exit(1)
    export(args.source)
